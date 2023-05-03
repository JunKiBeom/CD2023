import openpyxl


def find_addr(address):
    # 엑셀 파일 열기
    wb = openpyxl.load_workbook('mapdata.xlsx')

    try:
        # 주소를 띄어쓰기로 분할
        words = address.split()
        if words[0] == '충남':
            words[0] = '충청남도'
        elif words[0] == '충북':
            words[0] = '충청북도'
        elif words[0] == '경남':
            words[0] = '경상남도'
        elif words[0] == '경북':
            words[0] = '경상북도'
        elif words[0] == '전남':
            words[0] = '전라남도'
        elif words[0] == '전북':
            words[0] = '전라북도'


        # 첫 두 글자를 이용하여 시트 선택
        sheet_name = None
        first_two_letters = words[0][:2]

        for sheet in wb.sheetnames:
            if first_two_letters in sheet:
                sheet_name = sheet
                break

        if sheet_name is None:
            return None

        sheet = wb[sheet_name]

        # 단어들을 포함하는 행 찾기
        matching_rows = []
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
            if words[1] in str(row) and words[1] in str(row) and words[2] in str(row):
                if len(words) > 3:
                    if words[3] in str(row):
                        matching_rows.append(row)
                else:
                    matching_rows.append(row)

        if not matching_rows:
            return None

        # 위도와 경도 값을 찾아 반환
        coordinates = []
        for row in matching_rows:
            coordinate = (row[5], row[6])  # (F열 값, G열 값)
            coordinates.append(coordinate)

        return coordinates

    except KeyError:
        # 시트가 존재하지 않는 경우 None 반환
        return None

    finally:
        # 엑셀 파일 닫기
        wb.close()


# # 테스트 예시
# # address = "부산광역시 기장군 기장읍 내리"
# address = input()
# result = find_coordinates_in_excel(address)
# print(result)
#
# if result is not None:
#     for coordinate in result:
#         url = f"https://new.land.naver.com/complexes?ms={coordinate[0]},{coordinate[1]},15&a=APT:JGC:ABYG:OPST:PRE&b=B1:B2:B3&e=RETAIL"
#         print(url)
# #    print("주소에 해당하는 위도와 경도 값:")
# #    for coordinate in result:
# #        print(f"위도: {coordinate[0]}, 경도: {coordinate[1]}")
# else:
#     print("데이터를 찾을 수 없습니다.")
